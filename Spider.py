#!/bin/python3
import datetime
import json
import math
import os
import platform
import re
import sys
import threading
from urllib.parse import urlparse

import xlrd

import util
from openpyxl import load_workbook, Workbook
import docx


class Spider:
    conf = ""
    src_file_name = ""
    data = []
    data_res = []
    data_res_f = []

    def __init__(self, conf_fp=None):
        if conf_fp is None:
            conf_fp = "./conf/config.json"
        self.conf = util.read_conf(conf_fp)

        self.conf["g_src_dir"] = os.path.join(self.conf["data_dir"], self.conf["src_dir"])
        self.conf["g_dst_dir"] = os.path.join(self.conf["data_dir"], self.conf["dst_dir"])
        (self.src_file_name, self.abs_src_file_name) = util.select_file(self.conf["g_src_dir"])
        if self.src_file_name is None:
            return
        else:
            print(self.abs_src_file_name)
        self.conf["f_dst_dir"] = os.path.join(self.conf["g_dst_dir"], os.path.splitext(self.src_file_name)[0])
        self.conf["dst_file"] = os.path.join(self.conf["f_dst_dir"],
                                             "核_" + os.path.splitext(self.src_file_name)[0] + ".xlsx")
        if not os.path.exists(self.conf["data_dir"]):
            os.mkdir(self.conf["data_dir"])
        if not os.path.exists(self.conf["g_dst_dir"]):
            os.mkdir(self.conf["g_dst_dir"])
        if not os.path.exists(self.conf["f_dst_dir"]):
            os.mkdir(self.conf["f_dst_dir"])
        if self.conf["SCREEN"]:
            self.conf["screenshot_dir"] = os.path.join(self.conf["f_dst_dir"], "screenshot")
            if not os.path.exists(self.conf["screenshot_dir"]):
                os.mkdir(self.conf["screenshot_dir"])
        else:
            self.conf["screenshot_dir"] = None
        if self.conf["SCREEN"]:
            self.conf["pagesource_dir"] = os.path.join(self.conf["f_dst_dir"], "pagesource")
            if not os.path.exists(self.conf["pagesource_dir"]):
                os.mkdir(self.conf["pagesource_dir"])
        else:
            self.conf["pagesource_dir"] = None

    def init_data(self):
        if platform.system() == "Windows":
            os.system("ipconfig/flushdns")
        import shutil
        if self.data.__len__() > 0:
            return
        if os.path.splitext(self.src_file_name)[1] in ['.txt']:
            file_object = open(self.abs_src_file_name, 'r', encoding='utf-8')
            lines = file_object.readlines()
            file_object.close()
            for line in lines:
                new_line = line.replace('\n', '')
                match_obj = re.search(
                    r"([hH][tT]{2}[pP]://|[hH][tT]{2}[pP][sS]://|[wW]{3}.|[wW][aA][pP].|[fF][tT][pP].|[fF][iI][lL][eE].)[-A-Za-z0-9+&@#/%?=~_|!:,.;]+[-A-Za-z0-9+&@#/%=~_|]",
                    new_line)
                if match_obj:
                    row_t = [self.data.__len__() + 1, self.data.__len__() + 1, match_obj.group()]
                    self.data.append(row_t)
        elif os.path.splitext(self.src_file_name)[1] in ['.xlsx']:
            wb = load_workbook(self.abs_src_file_name)
            ws = wb.worksheets[0]
            for r in range(1, ws.max_row + 1):
                row_t = [r]
                for i in range(1, 3):
                    if ws.cell(row=r, column=i).value:
                        row_t.append(ws.cell(row=r, column=i).value)
                        row_t[1] = int(row_t[1])
                if row_t.__len__() == 3 and row_t[1] and row_t[2]:
                    self.data.append(row_t)
        elif os.path.splitext(self.src_file_name)[1] in ['.xls']:
            book = xlrd.open_workbook(self.abs_src_file_name)
            ws = book.sheet_by_index(0)
            for r in range(0, ws.nrows):
                row_t = [r+1]
                for i in range(0, 2):
                    row_t.append(ws.cell(r, i).value)
                    row_t[1] = int(row_t[1])
                # print(row_t)
                self.data.append(row_t)
        elif os.path.splitext(self.src_file_name)[1] in ['.docx']:
            d = docx.opendocx(self.abs_src_file_name)
            doc = docx.getdocumenttext(d)
            for line in doc:
                new_line = line.replace('\n', '')
                match_obj = re.search(r"https?://(?:[-\w.]|(?:%[\da-fA-F]{2}))+", new_line)
                if match_obj:
                    row_t = [self.data.__len__() + 1, self.data.__len__() + 1, match_obj.group()]
                    self.data.append(row_t)
        else:
            print("输入文件错误，系统不支持...")

        for i in self.data:
            (r_code, f_url) = util.get_url_normalize(i[2], short_urls=self.conf["short_domain"])  # formal URL
            if r_code == 404:
                (r_code, f_url) = util.get_url_normalize(i[2], short_urls=self.conf["short_domain"])
            i.append(f_url)
            if f_url == "异常":
                i.append("异常")
            else:
                i.append(urlparse((i[3])).hostname)
            len_t = 11
            if i[3] == '异常':
                for ii in range(0, len_t):
                    i.append('异常')
            else:
                for ii in range(0, len_t):
                    i.append('')
            if r_code >0:
                i[14] = r_code
        if os.path.splitext(self.src_file_name)[1] in ['.xlsx']:
            if not os.path.exists(self.conf["dst_file"]):
                shutil.copyfile(self.abs_src_file_name, self.conf["dst_file"])
                wb = load_workbook(self.conf["dst_file"])
                wb.save(self.conf["dst_file"])
        else:
            wb = Workbook()
            if os.path.splitext(self.src_file_name)[1] in ['.xlsx']:
                wb_s = load_workbook(self.abs_src_file_name)
                for sheet in wb_s.sheetnames:
                    ws_s = wb_s[sheet]
                    ws = wb[sheet]
                    for i, row in enumerate(ws_s.iter_rows()):
                        for j, cell in enumerate(row):
                            ws.cell(row=i + 1, column=j + 1, value=cell.value)
            wb.save(self.conf["dst_file"])
        print("*****获取网址：%d 条" % self.data.__len__())

    def show_conf(self):
        print(json.dumps(self.conf, sort_keys=True, indent=4, separators=(',', ':')))

    def show_data(self):
        print("SIZE: ", self.data.__len__(), len(self.data[0]))
        print('\n'.join(str(i) for i in self.data))

    def do_all(self, data=None):
        if data is None:
            data = self.data
        if self.conf["LOCATION"]:
            # print("\tDNS解析中:", threading.currentThread().ident)
            # data, conf=None, ipfix=False, headless=True, timeout=10
            data = util.do_dns(data=data, conf=self.conf)
        if self.conf["SCREEN"]:
            # print("\t浏览器访问中", threading.currentThread().ident)
            data = util.do_web(data=data, conf=self.conf)
        if self.conf["STATUS_CODE"]:
            # print("\t状态码获取中", threading.currentThread().ident)
            data = util.do_status(data=data)
        self.data_res.extend(data)

    def do_all_multi(self, data=None):
        if data is None:
            data = self.data
        threads = []
        n = math.ceil(data.__len__() / self.conf["poll"])
        dataSet = util.list_split(data, n)
        dataSetList = []
        for i in dataSet:
            dataSetList.append(i)
        for i in dataSetList:
            t = threading.Thread(target=self.do_all, args=(i,))
            t.start()
            threads.append(t)
        for thread in threads:
            thread.join()
        return self.data_res

    # 统计
    def statistics_data(self, tcoast=""):
        da_list = []
        da_statics = {"task": self.src_file_name, "总共": self.data.__len__(),
                      "正常": 0, "异常": 0, "境内": 0, "境外": 0, "跳转": 0,
                      "重点": 0, "非重点": 0, "打开": 0, "失败": 0, "耗时": tcoast, "进程数":self.conf["poll"]}
        for value in self.data_res_f:
            # print(len(value), value)
            da_t = [value[0], value[1], value[3], value[6], value[12], value[13]]
            if value[3] == "异常":
                da_statics["异常"] = da_statics["异常"] + 1
            else:
                da_statics["正常"] = da_statics["正常"] + 1
            if value[6] == "境内":
                da_statics["境内"] = da_statics["境内"] + 1
            elif value[6] == "境外":
                da_statics["境外"] = da_statics["境外"] + 1
            if value[10] == "是":
                da_statics["跳转"] = da_statics["跳转"] + 1
            if str(value[12]).isdigit() and int(value[12]) < 20000:
                da_statics["重点"] = da_statics["重点"] + 1
            else:
                da_statics["非重点"] = da_statics["非重点"] + 1
            if value[13] == "是" or (value[14] in self.conf["success_code"]):
                da_statics["打开"] = da_statics["打开"] + 1
            elif value[13] != "异常":
                da_statics["失败"] = da_statics["失败"] + 1
            da_list.append(da_t)
        print("结果汇总：", da_statics)
        if self.conf["all_records"]:
            util.update_task_excel(self.data_res_f, self.conf["dst_file"], sheet_name='全量结果', title="all")
        util.update_task_excel(da_list, self.conf["dst_file"], sheet_name='核查结果', title="smp")
        util.update_tj_excel(da_statics, os.path.join(self.conf["g_dst_dir"], self.conf["tongji"]))

    def run(self):
        if self.src_file_name is None:
            return
        # retry
        start = datetime.datetime.now()
        self.init_data()
        if self.data.__len__() == 0:
            return
        results_Retry = self.data
        for i in range(1, self.conf["run_counts"] + 1):
            print("*****运行第 %d 次, 待处理网址：%d" % (i, results_Retry.__len__()))
            if results_Retry.__len__() == 0:
                continue
            self.data_res = []
            self.do_all_multi(results_Retry)
            # 更新重新测试数据
            results_t = self.data_res
            results_OK = []
            results_Retry = []
            for value in results_t:
                if value[3] != "异常" and (value[5] == "异常" or value[6] == "境内" or value[13] == "否"):
                    if util.isIP(value[4]):
                        results_OK.append(value)
                    else:
                        results_Retry.append(value)
                else:
                    results_OK.append(value)
            self.data_res_f.extend(results_OK)
        self.data_res_f.extend(results_Retry)
        self.data_res_f = util.do_alexa(data=self.data_res_f, conf=self.conf)
        end = datetime.datetime.now()
        tt = str(datetime.timedelta(seconds=(end - start).seconds))
        self.statistics_data(tcoast=tt)
        print("====================\n%s任务结束，耗时 %s, \n拨测结果：%d条" % (end.strftime("%Y-%m-%d %H:%M:%S"), tt, len(self.data)))
        print("结果保存目录:", os.path.join(os.getcwd(), self.conf["dst_file"]).replace("./", ""))
        if platform.system() == "Windows":
            os.system("pause")


def main():
    # [2, 2, 'www.baidu.com', 'http://www.baidu.com', 'www.baidu.com', ['39.156.66.14', '39.156.66.18'], '境内',
    # '中国·北京', '百度一下，你就知道', 'https://www.baidu.com/', '是', 'https://www.baidu.com', '', '是', '', 1]
    # pyinstaller -F main.py

    # 初始化待测试数据
    spd = Spider(conf_fp=None)
    spd.run()


if __name__ == "__main__":
    main()
