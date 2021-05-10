#!/bin/python3
import datetime
import sys
from datetime import datetime
import json
import math
import os
import platform
import random
import re
import sqlite3
import time
from selenium.webdriver.support import expected_conditions as EC
from urllib.parse import urlparse
import dns.resolver
import ipdb
import requests
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side, colors
from requests.adapters import HTTPAdapter
from selenium import webdriver
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.webdriver.common.by import By

# 选择源文件
from selenium.webdriver.support.wait import WebDriverWait
from tldextract import tldextract
import traceback
import IPy


def select_file(fp):
    import shutil
    print("支持文件格式[.txt|.xlsx|.xls|.docx]")
    if platform.system() == "Windows":
        os.system("ipconfig/flushdns")
        from tkinter import Tk
        from tkinter import filedialog
        root = Tk()
        root.withdraw()
        filename = filedialog.askopenfilename(initialdir=fp)
        if not filename:
            print("选择文件失败，程序结束")
            return None, None
        return os.path.basename(filename), filename
    else:
        files = os.listdir(fp)
        files_set = []
        print("文件列表：")
        for file in files:
            if os.path.splitext(file)[1] in [".xlsx", ".xls", '.docx', '.txt']:
                files_set.append(file)
                print('[', len(files_set), ']:', file)
        index = input("支持文件格式[.txt|.xlsx|.xls|.docx]，请输入对应文件的序号:")
        file_name = files_set[int(index) - 1]
        dst_file = file_name
        return dst_file, os.path.join(fp, dst_file)


# 读取基本配置 in json format
def read_conf(conf_path):
    with open(conf_path, encoding='utf-8-sig') as f:
        data = json.load(f)
    return data


# 判断是否为IP
def isIP(str):
    p = re.compile('^((25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(25[0-5]|2[0-4]\d|[01]?\d\d?)$')
    if p.match(str):
        return True
    else:
        return False


# 分割列表
def list_split(listTemp, n):
    for i in range(0, len(listTemp), n):
        yield listTemp[i:i + n]


# 获取状态码和实际url（如发生跳转，返回跳转后的新url）
def get_status(url, quiet=False, logf=None):
    import urllib3
    urllib3.disable_warnings()
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.141 Safari/537.36'
    }
    s = requests.Session()
    s.mount('http://', HTTPAdapter(max_retries=3))
    s.mount('https://', HTTPAdapter(max_retries=3))
    try:
        # 最多n次跳转
        n = 3
        while n > 0:
            n = n - 1
            r = s.get(url=url, headers=headers, verify=False, timeout=(10, 5), allow_redirects=False)
            if not quiet:
                print(url, "\t:", r.status_code)
            print(r.headers)
            if r.status_code in [301, 302]:
                url = r.headers["Location"]
            else:
                s.close()
                return r.status_code, "SUCCESS", url
    except requests.exceptions.ConnectTimeout as e:
        # with open("./data/dst/log.txt" if logf is None else logf, "a") as f:
        #     traceback.print_exc(file=f)
        if not quiet:
            print("ConnectTimeout:\t", url)
        return e.errno, "ConnectTimeout", url
    except requests.exceptions.ReadTimeout as e:
        # with open("./data/dst/log.txt" if logf is None else logf, "a") as f:
        #     traceback.print_exc(file=f)
        if not quiet:
            print("ReadTimeout:\t", url)
        return e.errno, "ReadTimeout", url
    except requests.exceptions.ConnectionError as e:
        # with open("./data/dst/log.txt" if logf is None else logf, "a") as f:
        #     traceback.print_exc(file=f)
        if not quiet:
            print("ConnectionError:\t", url)
        return e.errno, "ConnectionError", url
    except requests.exceptions.HTTPError as e:
        # with open("./data/dst/log.txt" if logf is None else logf, "a") as f:
        #     traceback.print_exc(file=f)
        if not quiet:
            print("HTTPError:\t", url)
        return e.errno, "HTTPError", url
    except requests.exceptions.ConnectionResetError as e:
        # with open("./data/dst/log.txt" if logf is None else logf, "a") as f:
        #     traceback.print_exc(file=f)
        if not quiet:
            print("ConnectionResetError:\t", url)
        return e.errno, "ConnectionResetError", url
    except Exception as e:
        # with open("./data/dst/log.txt" if logf is None else logf, "a") as f:
        #     traceback.print_exc(file=f)
        if not quiet:
            print("未知错误:\t", url, e.__str__())
        s.close()
        return 500, "未知错误", url


# 标准化网址，return: status_code, new_url
def get_url_normalize(url, short_urls=None):
    url_new = url.replace(" ", "").replace('\n', "")
    url_new = url_new.split("\\")[0]
    default_scheme = "http"
    if urlparse(url).scheme == '':
        url_new = default_scheme + "://" + url
    if urlparse(url_new).hostname is None:
        url_new = "异常"
    elif "." not in urlparse(url_new).hostname:
        url_new = "异常"
    elif urlparse(url_new).hostname[0] == ".":
        url_new = "异常"
    elif len(urlparse(url_new).hostname) < 4:
        url_new = "异常"
    else:
        {}
    if url_new != "异常":
        # return 200, url_new
        domain = urlparse(url_new).hostname
        # if domain in short_urls:
        #     # print("短域名 ", url_new)
        #     (code, status, r_url) = get_status(url_new, quiet=True)
        #     # print((code, status, r_url))
        #     return code, r_url
        #     # print("\t跳转网址：", url_new)
        # elif len(domain) < 2:
        #     print("疑似短域名 ", domain)
        #     (code, status, r_url) = get_status(url_new, quiet=True)
        #     # print((code, status, r_url))
        #     return code, r_url
        # else:
        #     return -1, url_new
        return -1, url_new
    else:
        return 500, url_new


# 解析IP地址，返回状态码，以及对应的IP列表
def get_ips(domain, default_servers=None, quiet=False, logf=None):
    ips = []
    myResolver = dns.resolver.Resolver()
    if default_servers:
        myResolver.nameservers = default_servers
    try:
        A = myResolver.resolve(domain, "A")
        for i in A.response.answer:
            for j in i.items:
                if j.rdtype == 1:
                    ips.append(j.address)
        if not quiet:
            print("IPS: ", domain, ips)
        return 1, list(set(ips))
    except dns.resolver.NXDOMAIN:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        with open("./data/dst/log.txt" if logf is None else logf, "a") as f:
            traceback.print_exception(exc_type, exc_value, exc_traceback,
                                      limit=2, file=f)
        return 2, None
    except dns.resolver.NoNameservers:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        with open("./data/dst/log.txt" if logf is None else logf, "a") as f:
            traceback.print_exception(exc_type, exc_value, exc_traceback,
                                      limit=2, file=f)
        return 3, None
    except dns.resolver.NoAnswer:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        with open("./data/dst/log.txt" if logf is None else logf, "a") as f:
            traceback.print_exception(exc_type, exc_value, exc_traceback,
                                      limit=2, file=f)
        return 4, None
    except dns.exception.Timeout:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        with open("./data/dst/log.txt" if logf is None else logf, "a") as f:
            traceback.print_exception(exc_type, exc_value, exc_traceback,
                                      limit=2, file=f)
        return 5, None


# 初始化diver chrome
def init_driver(conf=None, PC=True):
    if conf is None:
        return
    if platform.system() == "Windows":
        chrome_fp = os.path.join(conf["conf"], "chromedriver.exe")
    else:
        chrome_fp = os.path.join(conf["conf"], "chromedriver")
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    chrome_options = Options()
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    if not PC:
        mobileEmulation = {'deviceName': 'iPhone X'}
        chrome_options.add_experimental_option('mobileEmulation', mobileEmulation)
    chrome_options.headless = conf["head_less"]
    chrome_options.add_argument(
        'user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36')
    service = Service(chrome_fp)
    driver = webdriver.Chrome(options=chrome_options, service=service)
    with open('./conf/stealth.min.js') as f:
        js = f.read()
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": js
    })
    df_timeout = conf["timeout_s"]
    driver.set_page_load_timeout(df_timeout)
    driver.set_script_timeout(df_timeout)
    return driver


# def init_driver_2(conf=None, PC=True):
#     from selenium import webdriver
#     from selenium.webdriver.chrome.service import Service
#     from webdriver_manager.chrome import ChromeDriverManager
#     from selenium.webdriver.common.by import By
#     from selenium.webdriver.chrome.options import Options
#     from selenium.webdriver.chrome.service import Service
#
#     chrome_options = Options()
#     chrome_options.add_argument('--ignore-certificate-errors')
#     chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
#     chrome_options.add_argument("--headless")
#     s = Service(ChromeDriverManager().install())
#     driver = webdriver.Chrome(service=s, options=chrome_options)
#     return driver


# 修订IP地址，ip138，返回：境内外，详情
def ip_fix(driver, ipstr, locat=None, detail=None, ip138=None, ipipnet=None, logf=None):
    try:
        driver.get("https://ip138.com/iplookup.asp?ip=%s&action=2" % ipstr)
        driver.find_element(By.ID, "ip").clear()
        driver.find_element(By.ID, "ip").send_keys(ipstr)
        driver.find_element(By.CLASS_NAME, "input-button").click()
        matchObj = re.search(r'[{]"ASN归属地"(.?)*[}]', driver.page_source, re.M | re.I)
        if matchObj:
            strinfo = json.loads(matchObj.group())
            if strinfo["ip_c_list"][0]['ct'] == "中国":
                if strinfo["ip_c_list"][0]['prov'].replace('特别行政区', '') in "中国,香港,澳门,台湾":
                    locat = "境外"
                    detail = strinfo["ip_c_list"][0]['ct'] + "·" + strinfo["ip_c_list"][0]['prov']
                    # print("\t更新IP位置:", strinfo['ASN归属地'])
            else:
                locat = "境外"
                detail = strinfo["ip_c_list"][0]['ct'] + "·" + strinfo["ip_c_list"][0]['prov']
        # else:
        # print("\t%s fix 失败, %s %s" % (ipstr, locat, detail))
    except Exception as e:
        # print("异常，当前函数：%s，%s" % ("ip_fix", e.__str__()))
        exc_type, exc_value, exc_traceback = sys.exc_info()
        with open("./data/dst/log.txt" if logf is None else logf, "a") as f:
            traceback.print_exception(exc_type, exc_value, exc_traceback,
                                      limit=2, file=f)
    return locat, detail


# 批量解析domain
def do_dns(data, conf=None):
    # print("\tDNS解析...")
    # data: 格式化网址列表
    if data is None:
        return
    if conf is None:
        conf = {}
        conf["conf"] = "./conf"
    if conf["ip_fix"]:
        driver = init_driver(conf=conf)
    data_res = []
    for value in data:
        if value[3] == "异常":
            continue
        if value[5] != '' and value[5] != "异常":
            continue
        domain = value[4]
        myaddr = []
        if type(domain) != str:
            print("value = ", value)
        if isIP(domain):
            myaddr.append(domain)
            value[5] = myaddr
            value[12] = "无"
        else:
            (retn, retv) = get_ips(domain, quiet=True, logf=os.path.join(conf["f_dst_dir"], conf["log"]))
            if retn == 1:
                value[5] = retv
            else:
                value[5] = "异常"  # IP
                value[6] = "境外"
                value[7] = "境外"
        if value[5] == "异常":
            continue
        dbpath = os.path.join(conf["conf"], "ipipfree.ipdb")
        dbc = ipdb.City(dbpath)
        city_str = dbc.find(value[5][0], "CN")
        jing_nei = "中国,香港,澳门,台湾"
        if city_str[0] in jing_nei:
            if city_str[1] in jing_nei:
                value[6] = "境外"
            else:
                value[6] = "境内"
        else:
            value[6] = "境外"
        value[7] = (city_str[0] + "·" + city_str[1])
        if conf["ip_fix"]:
            if value[6] == "境内":
                (a, b) = ip_fix(driver, value[5][0], locat=value[6], detail=value[7],
                                logf=os.path.join(conf["f_dst_dir"], conf["log"]))
                value[6] = a
                value[7] = b
            time.sleep(random.uniform(1, 3))
        # if re.match("127\.", value[5][0]):
        #     print(value[5][0])
        #     print("匹配")
        # else:
        #     print("未命中")
        if re.match("192\.168\.", value[5][0]) or value[5][0] in IPy.IP('172.16.0.0/12') or re.match("10\.", value[5][0]) or value[5][0] in ['0.0.0.0', '255.255.255.255'] or re.match("127\.", value[5][0]) or re.match("224\.", value[5][0]) or re.match("240\.", value[5][0]):
            value[6] = "境内"

    if conf["ip_fix"]:
        driver.quit()
    return data


# 访问失败时保存图片
def save_string_pic(conf, screen, strmsg, d_size):
    img = Image.new('RGB', (d_size), (255, 255, 255))
    img.save(screen)
    img = Image.open(screen)
    draw = ImageDraw.Draw(img)
    font_info = os.path.join(conf, "SIMLI.TTF")
    ttfont = ImageFont.truetype(font=font_info, size=40*round(d_size[0]/1000))
    draw.text((d_size[0]/4, d_size[1]/2), strmsg, fill="#0000ff", font=ttfont)
    img.save(screen)


# 单次浏览器访问网页获取内容
def do_url(conf, driver, url, screen=None, pagesrc=None, logf=None):
    # print("*****正在处理：", url)
    title = None
    current_url = url[3]
    ERRORSTR = {0: "正常",
                1: "net::ERR_CONNECTION_RESET",
                2: "net::ERR_NAME_NOT_RESOLVED",
                3: "net::ERR_CONNECTION_REFUSED",
                4: "Timeout",
                5: "Timed out receiving message from renderer",
                6: "net::ERR_CONNECTION_CLOSED",
                7: "未知"}
    ERRORNUM = 0  # 0:normal 1:连接重置 2:DNS解析失败
    try:
        driver.get(url[3])
        title = driver.title
        if title is None or title == "":
            title = url[3] + " real empty"
        current_url = driver.current_url
        # driver.execute_script('window.stop ? window.stop() : document.execCommand("Stop");')
        driver.save_screenshot(screen)
        source = driver.page_source
    except UnexpectedAlertPresentException as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        with open("./data/dst/log.txt" if logf is None else logf, "a") as f:
            traceback.print_exception(exc_type, exc_value, exc_traceback,
                                      limit=2, file=f, url=url[3])
        # print("\t弹窗:", url[3], e.__str__()[0:80])
        while True:
            try:
                WebDriverWait(driver, 10, 0.5).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
            except Exception as e:
                if "unexpected alert open" in e.__str__():
                    print("in UnexpectedAlertPresentException:", url[3], e.__str__())
                else:
                    title = driver.title
                    current_url = driver.current_url
                break
        if title is None or title == "":
            title = url[3] + " real empty"
        # driver.execute_script('window.stop ? window.stop() : document.execCommand("Stop");')
        driver.save_screenshot(screen)
        source = driver.page_source
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        try:
            with open("./data/dst/log.txt" if logf is None else logf, "a") as f:
                traceback.print_exception(exc_type, exc_value, exc_traceback,
                                          limit=2, file=f, url=url[3])
        except Exception as ee:
            print(ee.__str__())
        # with open("./data/dst/log.txt" if logf is None else logf, "a") as f:
        #     traceback.print_exception(exc_type, exc_value, exc_traceback,
        #                               limit=2, file=f, url=url[3])

        # print(e.__str__()[0:70])
        flag = False
        for er in conf["ERRORSTR"]:
            ERRORNUM = ERRORNUM + 1
            if er in e.__str__():
                flag = True
                if "Timed out receiving message from renderer" in e.__str__():
                    ERRORNUM = 0
                    title = url[3]
                    driver.execute_script('window.stop ? window.stop() : document.execCommand("Stop");')
                    time.sleep(2)
                    source = conf["ERRORSTR"][0]
                    try:
                        driver.save_screenshot(screen)
                    except Exception as e:
                        print("截图失败", url[2], e.__str__()[0:80])
                else:
                    title = "异常"
                    driver.get_window_size()
                    save_string_pic(conf["conf"], screen, conf["ERRORSTR"][ERRORNUM-1],
                                    (driver.get_window_size()['width'], driver.get_window_size()['height']))
                    source = conf["ERRORSTR"][ERRORNUM-1]
                # print("已知：", source)
                break
        if not flag:
            source = "未知"
            print("未知错误*****获取标题失败:", url[1], ":", ERRORNUM, url[2], e.__str__()[0:80])

        # if "net::ERR_CONNECTION_RESET" in e.__str__():
        #     ERRORNUM = 1
        # elif "net::ERR_NAME_NOT_RESOLVED" in e.__str__():
        #     ERRORNUM = 2
        # elif "net::ERR_CONNECTION_REFUSED" in e.__str__():
        #     ERRORNUM = 3
        # elif "net::ERR_CONNECTION_TIMED_OUT" in e.__str__():
        #     ERRORNUM = 4
        # elif "Timed out receiving message from renderer" in e.__str__():
        #     driver.execute_script('window.stop ? window.stop() : document.execCommand("Stop");')
        #     time.sleep(2)
        #     # driver.refresh()
        #     ERRORNUM = 5
        # elif "net::ERR_CONNECTION_CLOSED" in e.__str__():
        #     ERRORNUM = 6
        # else:
        #     ERRORNUM = 7
        #     ERRORSTR[7] = e.__str__()[0:20]
        #     print("*****获取标题失败:", url[1], ":", ERRORNUM, url[2], e.__str__()[0:80])
        # print("\t异常:", url[3], e.__str__()[0:80])

        # current_url = url[3]
        # if ERRORNUM == 5:
        #     title = url[3]
        #     ERRORNUM = 0
        #     source = ERRORSTR[ERRORNUM]
        #     try:
        #         driver.save_screenshot(screen)
        #     except Exception as e:
        #         print("截图失败", e.__str__()[0:80])
        # else:
        #     title = "异常"
        #     driver.get_window_size()
        #     save_string_pic(conf, screen, ERRORSTR[ERRORNUM], (driver.get_window_size()['width'], driver.get_window_size()['height']))
        #     source = ERRORSTR[ERRORNUM]
    finally:
        with open(pagesrc, 'wb') as f:
            f.write(source.encode("utf-8", "ignore"))
        return title, current_url, ERRORNUM


# 多网址浏览器访问
def do_web(data=None, conf=None):
    # print("\tweb访问...")
    if data is None:
        return
    driver = init_driver(conf=conf, PC=conf["PC"])
    if conf["PC"] and conf["screensize"]:
        driver.set_window_size(conf["screensize"][0], conf["screensize"][1])
    for value in data:
        start = datetime.now()
        # print("\t处理前：", value)
        if value[3] == "异常":
            continue
        # if value[5] == "异常": # DNS异常，未必网页真异常
        #     continue
        if value[8] != "" and value[8] != "异常":
            continue
        screen_file = os.path.join(conf["screenshot_dir"], str(value[1]) + "_" + value[4] + ".png")
        if os.path.exists(screen_file) and os.path.getsize(screen_file) > 1000000:
            continue
        # else:
        #     if os.path.exists(screen_file):
        #         print("\t更新....", str(value[1]) + "_" + value[4])
        page_file = os.path.join(conf["pagesource_dir"], str(value[1]) + "_" + value[4] + ".html")
        (title, url, FD) = do_url(conf, driver, value, screen=screen_file, pagesrc=page_file,
                                  logf=os.path.join(conf["f_dst_dir"], conf["log"]))
        value[8] = title
        value[9] = url
        default_pro = urlparse(value[2]).scheme
        if default_pro == "":
            default_pro = "http"
        if (urlparse(url).scheme + "://" + urlparse(url).hostname) == \
                (default_pro + "://" + value[4]):
            value[10] = "否"
        else:
            value[10] = "是"
        if urlparse(url).scheme == "https":
            value[11] = urlparse(url).scheme + "://" + urlparse(url).hostname
        else:
            value[11] = url
        if FD == 0:
            value[13] = "是"
        else:
            value[13] = "否"
        # print("\t处理后：", value)
        end = datetime.now()
        # print("url time:", value[3], "\t", (end - start).seconds)
        value[15] = (end - start).seconds
    driver.quit()
    return data


# 获取排名-通过link114.cn
def do_domain_alexa_114(browser, da):
    # print("***** link114")
    url_114 = "http://www.link114.cn/alexa/"
    time.sleep(random.randint(1, 5))
    browser.get(url_114)
    col_l = 50
    yum = len(da) % col_l
    counts = math.ceil(len(da) / col_l)
    col_add = col_l
    for i in range(counts):
        if (i == (counts - 1)) and (yum > 0):
            col_add = yum
        domain_set = []
        for j in range(col_add):
            index = da[i * col_l + j][1]
            domain = da[i * col_l + j][4]
            # print("查询域名：", domain)
            if da[i * col_l + j][12] != "无":
                domain_set.append(domain)
        domain_str = '#,#'.join(domain_set)
        # logger.info(domain_str)
        browser.find_element(By.ID, "ip_websites").clear()
        browser.find_element(By.ID, "ip_websites").send_keys(domain_str)
        browser.find_element(By.ID, "tj").click()
        time.sleep(20)
        trlist = browser.find_elements(By.TAG_NAME, "tr")
        data_dict = {}
        index = 0
        for tr in trlist:
            # 获取tr中的所有td
            tdlist = tr.find_elements(By.TAG_NAME, "td")
            data_dict[index] = []
            if len(tdlist) > 0:
                # 获取td[0]的文本
                text_1 = tr.find_elements(By.TAG_NAME, "td")[0].text
                text_1 = text_1.replace(".", "")
                # trid_1 = tr.get_attribute("id")
                trid_1 = tr.find_elements(By.TAG_NAME, "td")[1].get_attribute("value")
                trid_alexa = tr.find_elements(By.TAG_NAME, "td")[2].text
                if trid_alexa is not None:
                    # print(trid_alexa)
                    if "无" in trid_alexa:
                        trid_alexa = "无"
                    if "重查" in trid_alexa:
                        trid_alexa = "无"
                    if "Alexa:" in trid_alexa:
                        trid_alexa = int(trid_alexa.replace("Alexa:", ""))
                    data_dict[index].append(text_1)
                    data_dict[index].append(trid_1)
                    data_dict[index].append(trid_alexa)
                    index = index + 1
        alexa_set = data_dict.values()
        for d in alexa_set:
            for value in da:
                if d[1] == value[4]:
                    value[12] = d[2]
    return da


# 获取排名
def do_alexa(conf, data=None):
    if data is None:
        return
    if not conf["ALEXA"]:
        return data
    print("\t查询排名中")
    alexa_yes = []
    alexa_no = []
    for value in data:
        if value[3] == "异常" or isIP(value[4]):
            value[12] = "异常" if value[3] == "异常" else "无"
            alexa_no.append(value)
        else:
            alexa_yes.append(value)
    alexa_way = conf["ONLINE"]
    # driver = init_driver(conf)
    c = None
    if conf["OFFLINE"]:
        conn = sqlite3.connect(os.path.join(conf["conf"], conf["LOCAL_DB"]))
        c = conn.cursor()
    if alexa_way == 1:  # offline + 重点查询
        file_path = "file:///" + os.path.abspath('.').replace('\\', '/') + "/conf/public_suffix_list.dat"
        no_fetch = tldextract.TLDExtract(suffix_list_urls=[file_path])
        for value in alexa_yes:
            alexa_d = "无"
            if value[12] != "" and value[12] != 0:
                continue
            if value[3] == "异常" or isIP(value[4]):
                value[12] = "无"
            else:
                if c:
                    sql = "select Alexa_COM from top2w where Domain=?"
                    # ext = tldextract.extract(value[4])
                    ext = no_fetch(value[4])
                    select_data = c.execute(sql, (ext.domain + "." + ext.suffix,))
                    for row in select_data:
                        print("本次排名库中找到...", ext.domain + "." + ext.suffix, row[0])
                        alexa_d = row[0]
                # else:
                #     alexa_d = do_domain_alexa_online(driver, value[4], self.conf["APIKEY"])
            # time.sleep(random.uniform(1, 4))
            value[12] = alexa_d
    elif alexa_way == 2:  # link 114
        driver = init_driver(conf=conf)
        alexa_yes = do_domain_alexa_114(driver, alexa_yes)
        driver.quit()
    if c:
        c.close()

    alexa_yes.extend(alexa_no)
    return alexa_yes


# 获取状态码
def do_status(data=None):
    if data is None:
        return
    for value in data:
        if value[3] == "异常":
            continue
        (code, status, r_url) = get_status(value[3], quiet=True)
        time.sleep(random.randint(0, 4))
        # print(code, ":", status, "\t", value[3])
        value[14] = code if code else 400
    return data


# 更新任务excel
def update_task_excel(da, filename, sheet_name, title=None):
    # print("update_task_excel...", str(da))
    wb = load_workbook(filename)
    ws_1 = wb.create_sheet(sheet_name, 0)
    wb.active = 0
    ws_1 = wb.worksheets[0]
    Color = ['ffffff', '000000']
    alig_s = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False)
    fille = PatternFill('solid', fgColor=Color[0])
    font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color=Color[1])
    border_set_H = Border(left=Side(style='medium', color=colors.BLACK),
                          right=Side(style='medium', color=colors.BLACK),
                          top=Side(style='medium', color=colors.BLUE),
                          bottom=Side(style='medium', color=colors.BLACK))
    border_set_T = Border(left=Side(style='thin', color=colors.BLACK),
                          right=Side(style='thin', color=colors.BLACK),
                          top=Side(style='thin', color=colors.BLACK),
                          bottom=Side(style='thin', color=colors.BLACK))

    smptitle = ['序号', '网址', '境内外', '排名', '打开参考']
    alltitle = ['序号', '源网址', '格式化', '域名', 'IP', '境内外', 'Location', '标题', '最终访问', '跳转', '提取域名', '排名', '打开参考', '状态码',
                '耗时']
    if title == "smp" or title is None:
        for i in range(len(smptitle)):
            ws_1.cell(row=1, column=i + 1).value = smptitle[i]
            ws_1.cell(row=1, column=i + 1).alignment = alig_s
            ws_1.cell(row=1, column=i + 1).font = font
            ws_1.cell(row=1, column=i + 1).fill = fille
            ws_1.cell(row=1, column=i + 1).border = border_set_H
    elif title == "all":
        for i in range(len(alltitle)):
            ws_1.cell(row=1, column=i + 1).value = alltitle[i]
            ws_1.cell(row=1, column=i + 1).alignment = alig_s
            ws_1.cell(row=1, column=i + 1).font = font
            ws_1.cell(row=1, column=i + 1).fill = fille
            ws_1.cell(row=1, column=i + 1).border = border_set_H
    index = 1
    for value in da:
        for ii in range(2, len(value) + 1):
            if '[' in str(value[ii - 1]):
                strvalue = '\n'.join(value[ii - 1]).replace('[', '').replace(']', '').replace('\'', '')
            else:
                strvalue = value[ii - 1]
            ws_1.cell(row=value[0] + 1, column=ii - 1).value = strvalue
            ws_1.cell(row=value[0] + 1, column=ii - 1).alignment = alig_s
            ws_1.cell(row=value[0] + 1, column=ii - 1).border = border_set_T

    wb.save(filename)


# 更新总汇总表
def update_tj_excel(data, filename):
    if not os.path.exists(filename):
        wb = Workbook()
    else:
        wb = load_workbook(filename)
    ws = wb.worksheets[0]
    Color = ['ffffff', '000000']
    alig_s = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False)
    fille = PatternFill('solid', fgColor=Color[0])
    font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color=Color[1])
    border_set_H = Border(left=Side(style='medium', color=colors.BLACK),
                          right=Side(style='medium', color=colors.BLACK),
                          top=Side(style='medium', color=colors.BLUE),
                          bottom=Side(style='medium', color=colors.BLACK))
    border_set_T = Border(left=Side(style='thin', color=colors.BLACK),
                          right=Side(style='thin', color=colors.BLACK),
                          top=Side(style='thin', color=colors.BLACK),
                          bottom=Side(style='thin', color=colors.BLACK))
    alig_s = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False)
    title = ["序号", "任务名", "网址数目", "正常", "异常",
             "境内", "境外", "跳转", "重点", "非重点", "打开", "打不开", "境内比例", "重点比例", "打不开比例", "耗时", "进程数", "日期"]
    if ws.max_row == 1:
        for i in range(len(title)):
            ws.cell(row=1, column=i + 1).value = title[i]
            ws.cell(row=1, column=i + 1).alignment = alig_s
            ws.cell(row=1, column=i + 1).font = font
            ws.cell(row=1, column=i + 1).fill = fille
            ws.cell(row=1, column=i + 1).border = border_set_H

    ws.cell(row=ws.max_row + 1, column=1).value = ws.max_row
    ws.cell(row=ws.max_row, column=2).value = data["task"]
    ws.cell(row=ws.max_row, column=3).value = data["总共"]
    ws.cell(row=ws.max_row, column=4).value = data["正常"]
    ws.cell(row=ws.max_row, column=5).value = data["异常"]
    ws.cell(row=ws.max_row, column=6).value = data["境内"]
    ws.cell(row=ws.max_row, column=7).value = data["境外"]
    ws.cell(row=ws.max_row, column=8).value = data["跳转"]
    ws.cell(row=ws.max_row, column=9).value = data["重点"]
    ws.cell(row=ws.max_row, column=10).value = data["非重点"]
    ws.cell(row=ws.max_row, column=11).value = data["打开"]
    ws.cell(row=ws.max_row, column=12).value = data["失败"]
    ws.cell(row=ws.max_row, column=13).value = '{:.1f}%'.format(data["境内"] * 100 / data["总共"])
    ws.cell(row=ws.max_row, column=14).value = '{:.1f}%'.format(data["重点"] * 100 / data["总共"])
    ws.cell(row=ws.max_row, column=15).value = '{:.1f}%'.format(data["失败"] * 100 / data["总共"])
    ws.cell(row=ws.max_row, column=16).value = data["耗时"]
    ws.cell(row=ws.max_row, column=17).value = data["进程数"]
    ws.cell(row=ws.max_row, column=18).value = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    wb.save(filename)


def test():
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    chrome_options = Options()
    # chrome_options.add_argument('--ignore-certificate-errors')
    # chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

    chrome_options.add_argument(
        'user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36')
    service = Service("./conf/chromedriver.exe")
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--headless')
    driver = webdriver.Chrome(options=chrome_options, service=service)
    # 打开百度
    driver.get("https://www.baidu.com/")
    driver.quit()


def main():
    # url = "http://www.baidu.com"
    # conf = {"conf": "./conf", "head_less": True, "timeout_s": 20}
    # driver = init_driver(conf=conf)
    # driver.get(url)
    # time.sleep(5)
    # driver.quit()
    test()


if __name__ == "__main__":
    main()
